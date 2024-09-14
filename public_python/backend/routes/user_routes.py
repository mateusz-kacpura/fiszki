## System tagowania zadaĹ„ i szybliego wyszukiwania interesujacych zadaĹ„ oparty o bardziej zaawansowanÄ… tabelÄ™ json
# System dodawania pomysĹ‚, co naleĹĽaĹ‚o by zmieniÄ‡ w aplikacji, jaki pomysĹ‚ wdroĹĽyÄ‡ oparty o nowÄ… tabelÄ™ json
# System dodawania informacji o znalezienionym issues
# System dodawania nowych zadaĹ„
# Nowy moduĹ‚ nauki polegajÄ…cy na Ĺ‚Ä…czeniu podobne sĹ‚owa w pary, naleĹĽy stwoĹĽyÄ‡ nowÄ… tabelÄ™ json
# MoĹĽliwoĹ›Ä‡ dodania wĹ‚asnego kodu javascript dla uĹĽytkownika, oraz kodu html wczytywanego z json uĹĽytkownika



# issues
# image_words_learning.js nie dziaĹ‚a odtwarzanie dzwieku po odwruceniu kierunku audio
# POST	http://127.0.0.1:5000/user/text-to-speech Error: Error: Failed to generate speech
# 
# ## 

from flask import Blueprint, render_template, request, jsonify, make_response
from flask_login import login_required, current_user
from flask import send_from_directory
from functools import partial
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from scipy.io.wavfile import write
import pandas as pd 
import json
import os

from backend.models.user_model import User
from backend.routes.users.learn.process_words import process_words
from backend.routes.users.learn.get_texts_for_insert_words_to_text import get_texts_for_insert_words_to_text
from backend.routes.users.learn.get_texts_data_for_insert_fot_text import get_texts_data_for_insert_fot_text
from backend.routes.users.learn.save_history_translations_for_insert_words_to_text import save_history_translations_for_insert_words_to_text
from backend.routes.users.learn.load_translation_history_for_insert_wors_to_text import load_translation_history_for_insert_wors_to_text
from backend.routes.users.AI.model_fb_translate import model_fb_translate
from backend.routes.users.modals.modal_pop_up_for_definition_learning import modal_pop_up_for_definition_learning
from backend.routes.users.modals.modal_pop_up_for_image_learning import modal_pop_up_for_image_learning
from backend.routes.users.modals.modal_pop_up_for_insert_word import modal_pop_up_for_insert_word
from backend.routes.users.modals.modal_pop_up_for_multi_learning import modal_pop_up_for_multi_learning
from backend.routes.users.AI.real_time_speech_recognition import real_time_speech_recognition
from backend.routes.users.AI.text_to_speech import text_to_speech
from backend.routes.users.AI.text_to_speech_groq import text_to_speech_groq # taki model w groq obecnie nie istnieje
from backend.routes.users.manage.load_audio_paths import load_audio_paths 
from backend.routes.users.manage.load_image_paths import load_image_paths
from backend.routes.users.manage.download_configuration import download_configuration
from backend.routes.users.save_statistic import save_statistic
from backend.routes.users.save_setting import save_setting
from backend.routes.users.manage.uploads_save_json import uploads_save_json
from backend.routes.users.manage.upload_excel import upload_excel

user_route = Blueprint('user', __name__, url_prefix='/user', template_folder='templates')
user_route.route('/process-words', methods=['POST'])(process_words)
user_route.route('/get_texts_for_insert_words_to_text')(get_texts_for_insert_words_to_text)
user_route.route('/get_texts_data_for_insert_fot_text')(get_texts_data_for_insert_fot_text)
user_route.route('/save_history_translations_for_insert_words_to_text', methods=['POST'])(save_history_translations_for_insert_words_to_text)
user_route.route('/load_translation_history_for_insert_wors_to_text', methods=['GET'])(load_translation_history_for_insert_wors_to_text)
user_route.route('/model_fb_translate', methods=['POST'])(model_fb_translate)
user_route.route('/modals/modal_pop_up_for_definition_learning', methods=['GET'])(modal_pop_up_for_definition_learning)
user_route.route('/modals/modal_pop_up_for_image_learning', methods=['GET'])(modal_pop_up_for_image_learning)
user_route.route('/modals/modal_pop_up_for_insert_word', methods=['GET'])(modal_pop_up_for_insert_word)
user_route.route('/modals/modal_pop_up_for_multi_learning', methods=['GET'])(modal_pop_up_for_multi_learning)


# funkcja nie dziaĹ‚a w peĹ‚ni offline pomimo pobrania repozytoriĂłw whisper do folderu models
user_route.route('/real-time-speech-recognition', methods=['POST'])(real_time_speech_recognition)
user_route.route('/text-to-speech', methods=['POST'])(text_to_speech)
user_route.route('/text-to-speech-groq', methods=['POST'])(text_to_speech_groq) # taki model w groq obecnie nie istnieje

user_route.route('/load-audio-paths', methods=['POST'])(load_audio_paths)
user_route.route('/load-image-paths', methods=['POST'])(load_image_paths)
user_route.route('/download_configuration', methods=['POST'])(download_configuration)
user_route.route('/save_statistic', methods=['POST'])(save_statistic)
user_route.route('/saveSetting', methods=['POST'])(save_setting)
user_route.route('/uploads-save-json', methods=['POST'])(uploads_save_json)
user_route.route('/upload_excel', methods=['POST'])(upload_excel)

# Configurations
UPLOAD_FOLDER = 'uploads'
IMAGE_FOLDER = 'baza_danych/image_files/English/'
AUDIO_FOLDER = 'baza_danych/audio_files/English/'
LOG_FOLDER = 'logi'
STATISTICS_FILE ='baza_danych/statistic/statistics.json'
SETTING_FILE ='baza_danych/setting/excludedWords.json'

os.environ["CUDA_VISIBLE_DEVICES"] = "0"
os.environ["SUNO_USE_SMALL_MODELS"] = "1"

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(IMAGE_FOLDER):
    os.makedirs(IMAGE_FOLDER)
if not os.path.exists(AUDIO_FOLDER):
    os.makedirs(AUDIO_FOLDER)
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

for folder in [UPLOAD_FOLDER, IMAGE_FOLDER, AUDIO_FOLDER, LOG_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

routes = [
    ('/', 'index.html', None, 'index'),
    ('/learn', 'learn.html', None, 'learn'),
    ('/manage', 'manage.html', None, 'manage'),
    ('/learning/sentences_learning', 'learning/sentences_learning.html', None, 'sentences_learning'),
    ('/learning/multi_learning', 'learning/multi_learning.html', None, 'multilearning'),
    ('/learning/scattered_words_learning', 'learning/scattered_words_learning.html', None, 'scattered_words_learning'),
    ('/learning/insert_word', 'learning/insert_word.html', None, 'insert_word'),
    ('/learning/definition', 'learning/definition.html', None, 'definition'),
    ('/learning/single_word_learning', 'learning/single_word_learning.html', "Language Quiz", 'single_word_learning'),
    ('/learning/image_words_learning', 'learning/image_words_learning.html', "Image learning", 'image_words_learning'),
    ('/learning/insert_word_to_text', 'learning/insert_word_to_text.html', 'Insert word to text', 'insert_word_to_text'),
    ('/learning/synonims', 'learning/insert_synonims_to_text.html', 'Insert synonims to text', 'insert_synonims_to_text'),
    ('/manage/exel', 'manage/exel.html', 'Excel', 'exel'),
]

def create_view_function(template, title=None):
    def view_function():
        if title:
            return render_template(template, title=title)
        else:
            return render_template(template)
    return view_function

for route, template, title, endpoint in routes:
    view_func = partial(create_view_function(template, title))
    user_route.add_url_rule(route, view_func=view_func, endpoint=endpoint)

# Define route information
routes_info = [
    ('/baza_danuch/setting/<path:filename>', 'baza_danych/setting', None, 'get_excluded_words'),
    ('/learning/baza_danych/setting/<path:filename>', 'baza_danych/setting', None, 'get_excluded_words_learning'),
    ('/baza_danych/statistic/statistics.json', 'baza_danych/statistic', 'statistics.json', 'get_statistics'),
    ('/learning/image_files/English/<path:filename>', 'baza_danych/image_files/English', None, 'custom_static_images'),
    ('/learning/baza_danych/audio_files/English/<path:filename>', 'baza_danych/audio_files/English', None, 'custom_static_audio'),
    ('/image_files/English/<path:filename>', 'baza_danych/image_files/English', None, 'image_files'),
    ('/audio_files/English/<path:filename>', 'baza_danych/audio_files/English', None, 'audio_files'),
    ('/audio_files/English/<path:path>', AUDIO_FOLDER, None, 'send_audio')
]

@user_route.route('/<path:path>')
@login_required
def send_static(path):
    return send_from_directory('public', path)

# Register routes dynamically
for route, directory, filename, func_name in routes_info:
    def create_route_func(directory, filename):
        def route_func(filename):
            return send_from_directory(directory, filename)
        return route_func

    if filename:
        func = create_route_func(directory, filename)
    else:
        func = create_route_func(directory, None)

    func.__name__ = func_name
    user_route.route(route)(login_required(func))

@user_route.route('/get_synonim_data')
def get_synonym_data():
    with open('baza_danych/user_datas/test/synonim_data.json', 'r', encoding='utf-8') as f:
        synonym_data = json.load(f)
    return jsonify(synonym_data)

@user_route.route('/save', methods=['POST'])
@login_required
def save_json():
    file_name = request.json['fileName']
    json_data = request.json['jsonData']
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    with open(file_path, 'w') as f:
        json.dump(json_data, f, indent=2)
    print(f'Saved JSON file: {file_path}')  # Print detailed information
    return jsonify({'message': 'Data saved successfully as JSON'})

@user_route.route('/words', methods=['GET'])
@login_required
def get_words():
    file_name = request.args.get('file')
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            data = json.load(f)
            return jsonify(data)
    else:
        return jsonify([])

@user_route.route('/edit/:file/:index', methods=['POST'])
@login_required
def edit_word():
    file_name = request.args.get('file')
    index = int(request.args.get('index'))
    updated_data = request.json
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    if os.path.exists(file_path):
        with open(file_path, 'r+') as f:
            data = json.load(f)
            if index >= 0 and index < len(data):
                data[index] = updated_data
                f.seek(0)
                json.dump(data, f, indent=2)
                return jsonify({'message': 'Data updated successfully'})
            else:
                return jsonify({'error': 'Invalid index'})
    else:
        return jsonify({'error': 'File not found'})

@user_route.route('/files', methods=['GET'])
@login_required
def get_files():
    files = os.listdir(UPLOAD_FOLDER)
    return jsonify(files)

@user_route.route('/upload', methods=['POST'])
@login_required
def upload_file():
    file = request.files['file']
    file_path = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))
    file.save(file_path)
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return jsonify({'columns': data[0]})

@user_route.route('/data', methods=['GET'])
@login_required
def get_data():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 5))
    start = (page - 1) * per_page
    end = start + per_page
    return jsonify(data[start:end])

@user_route.route('/data', methods=['POST'])
@login_required
def add_data():
    new_item = request.json
    data.append(new_item)
    return jsonify(data)

@user_route.route('/data/<int:index>', methods=['PUT'])
@login_required
def update_data(index):
    updated_item = request.json
    data[index] = updated_item
    return jsonify(data)

@user_route.route('/data/<int:index>', methods=['DELETE'])
@login_required
def delete_data(index):
    del data[index]
    return jsonify(data)

# Globalna zmienna do przechowywania danych
data = []

@user_route.route('/parse_excel_columns', methods=['POST'])
@login_required
def parse_excel_columns():
    file = request.files['file']
    df = pd.read_excel(file)
    columns = df.columns.tolist()
    return jsonify(columns=columns)

@user_route.route('/upload_json', methods=['POST'])
@login_required
def upload_json():
    file = request.files['file']
    global data
    data = json.load(file)
    return jsonify(data)

@user_route.route('/data/count', methods=['GET'])
@login_required
def get_data_count():
    print("len data/count", len(data))
    return jsonify(len(data))

@user_route.route('/upload-save-json', methods=['POST'])
@login_required
def upload_save_json():
    global data
    data = request.json
    with open('data.json', 'w') as f:
        json.dump(data, f)
    return jsonify({"success": True})

@user_route.route('/translate-word', methods=['GET'])
@login_required
def translate_word():
    word = request.args.get('word')
    target_lang = request.args.get('targetLang')
    # translation = translate(word, target_lang)
    # return jsonify(translation)
