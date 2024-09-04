from flask import Blueprint, render_template, request, jsonify, make_response
from flask_login import UserMixin, login_required
from flask import send_from_directory
from models.user_model import User
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from transformers import AutoProcessor, AutoModel, BarkModel, WhisperProcessor, WhisperForConditionalGeneration
from transformers import pipeline
import torch
from scipy.io.wavfile import write
from pydub import AudioSegment
import nltk
from nltk.corpus import wordnet
from nltk.stem import WordNetLemmatizer
import pyaudio
import numpy as np
import scipy
import pandas as pd 
from io import BytesIO, StringIO
import csv
from groq import Groq
import json
import logging
import requests
import os

user_route = Blueprint('user', __name__, url_prefix='/user', template_folder='templates')

# Configurations
UPLOAD_FOLDER = 'uploads'
IMAGE_FOLDER = 'baza_danych/image_files/English/'
AUDIO_FOLDER = 'baza_danych/audio_files/English/'
LOG_FOLDER = 'logi'
STATISTICS_FILE ='baza_danych/statistic/statistics.json'
SETTING_FILE ='baza_danych/setting/excludedWords.json'

#MODELS 
BARK_MODEL = "models/files/bark" # git clone https://huggingface.co/suno/bark
WHISPER_MEDIUM = "models/files/whisper-medium" # git clone https://huggingface.co/openai/whisper-medium

os.environ["CUDA_VISIBLE_DEVICES"] = ""
os.environ["SUNO_USE_SMALL_MODELS"] = "1"

# Utwórzenie loggerów do logowania ogólnego
app_logger = logging.getLogger('app_logger')
app_logger.setLevel(logging.INFO)
app_handler = logging.FileHandler('app.log')
app_handler.setLevel(logging.INFO)
app_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
app_handler.setFormatter(app_formatter)
app_logger.addHandler(app_handler)

# Utwórzenie logger do logowania ścieżek audio
audio_logger = logging.getLogger('audio_logger')
audio_logger.setLevel(logging.INFO)
audio_handler = logging.FileHandler(os.path.join(LOG_FOLDER, 'path_audio.log'))
audio_handler.setLevel(logging.INFO)
audio_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
audio_handler.setFormatter(audio_formatter)
audio_logger.addHandler(audio_handler)

# Utwórzenie loggerów do logowania obrazów
image_logger = logging.getLogger('image_logger')
image_logger.setLevel(logging.INFO)
image_handler = logging.FileHandler(os.path.join(LOG_FOLDER, 'path_image.log'))
image_handler.setLevel(logging.INFO)
image_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
image_handler.setFormatter(image_formatter)
image_logger.addHandler(image_handler)

# Konfiguracja loggera do logowania plików
upload_logger = logging.getLogger('upload_logger')
upload_logger.setLevel(logging.INFO)
upload_handler = logging.FileHandler(os.path.join(LOG_FOLDER, 'path_uploads.log'))
upload_handler.setLevel(logging.INFO)
upload_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
upload_handler.setFormatter(upload_formatter)
upload_logger.addHandler(upload_handler)

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(IMAGE_FOLDER):
    os.makedirs(IMAGE_FOLDER)
if not os.path.exists(AUDIO_FOLDER):
    os.makedirs(AUDIO_FOLDER)
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)

for folder in [UPLOAD_FOLDER, IMAGE_FOLDER, AUDIO_FOLDER, LOG_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)


lemmatizer = WordNetLemmatizer()

def get_wordnet_pos(treebank_tag):
    """Converts treebank tags to wordnet tags."""
    if treebank_tag.startswith('J'):
        return wordnet.ADJ
    elif treebank_tag.startswith('V'):
        return wordnet.VERB
    elif treebank_tag.startswith('N'):
        return wordnet.NOUN
    elif treebank_tag.startswith('R'):
        return wordnet.ADV
    else:
        return wordnet.NOUN

def lemmatize_word(word):
    """Lemmatizes the given word using its part of speech tag."""
    pos_tag = nltk.pos_tag([word])[0][1]
    wordnet_pos = get_wordnet_pos(pos_tag)
    return lemmatizer.lemmatize(word, wordnet_pos)

@user_route.route('/process-words', methods=['POST'])
@login_required
def process_words():
    try:
        data = json.loads(request.data)
        words = data.get('words', [])
        
        lemmatized_words = [{**word, 'lemma': lemmatize_word(word['word'])} for word in words]
        
        return jsonify({'words': lemmatized_words})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@user_route.route('/')
@login_required
def index():
    return render_template('index.html')

@user_route.route('/learn')
@login_required
def learn():
    return render_template('learn.html')

@user_route.route('/manage')
@login_required
def manage():
    return render_template('manage.html')

@user_route.route('/learning/sentences_learning')
@login_required
def sentences_learning():
    return render_template('learning/sentences_learning.html')

@user_route.route('/learning/multi_learning')
@login_required
def multilearning():
    return render_template('learning/multi_learning.html')

@user_route.route('/learning/scattered_words_learning')
@login_required
def scattered_words_learning():
    return render_template('learning/scattered_words_learning.html')

@user_route.route('/learning/insert_word')
@login_required
def insert_word():
    return render_template('learning/insert_word.html')

@user_route.route('/learning/definition')
@login_required
def definition():
    return render_template('learning/definition.html')

@user_route.route('/learning/single_word_learning')
@login_required
def single_word_learning():
    return render_template('learning//single_word_learning.html', title="Language Quiz")

@user_route.route('/learning/image_words_learning')
@login_required
def image_words_learning():
    return render_template('learning/image_words_learning.html', title="Image learning")

@user_route.route('/get_text_data')
@login_required
def get_text_data():
    name = request.args.get('name')  # Pobieranie parametru 'name' z zapytania
    with open('baza_danych/user_datas/test/insert_word.json', encoding='utf-8') as f:
        text_data = json.load(f)

    # Jeśli name jest podane, znajdź odpowiedni tekst
    if name:
        selected_text = next((text for text in text_data["texts"] if text["name"] == name), None)
        if selected_text:
            return jsonify({"texts": [selected_text]})
    
    # Jeśli nie podano name, zwróć listę nazw dostępnych tekstów
    text_names = [{"name": text["name"], "uuid": text["uuid"]} for text in text_data["texts"]]
    return jsonify({"text_names": text_names})

# Route for rendering the HTML page
@user_route.route('/learning/insert_word_to_text')
@login_required
def insert_word_to_text():
    return render_template('learning/insert_word_to_text.html', title="Insert word to text")

@user_route.route('/manage/exel')
@login_required
def exel():
    return render_template('manage/exel.html', title="Excel")

@user_route.route('/translate', methods=['POST'])
@login_required
def translate():
    data = request.json
    text = data.get('text')

    if not text:
        return jsonify({'error': 'No text provided'}), 400

    try:
        # Initialize the translation pipeline
        generator = pipeline("translation", model="facebook/m2m100_418M")
        result = generator(text, src_lang="en", tgt_lang="pl", max_length=512)

        # Extract the translated text from the result
        translated_text = result[0]['translation_text']
        return jsonify({'translatedText': translated_text})

    except Exception as e:
        # Return error message and status code
        return jsonify({'error': str(e)}), 500

@user_route.route('/modals/definition-pop-up', methods=['GET'])
@login_required
def definition_pop_up():
    try:
        resultMessage = request.args.get('resultMessage', '')
        selectedWord = request.args.get('selectedWord', '')
        word = request.args.get('word', '')
        print(word)
        print(selectedWord)
        translation = request.args.get('translation', '')
        definition = request.args.get('definition', '')
        theme = request.args.get('theme', 'light')

        # Determine if the answer is correct
        isCorrect = resultMessage == 'Correct!'
        modalHeaderClass = 'bg-success text-white' if selectedWord == word else 'bg-danger text-white'
        modalTitle = 'Correct Answer!' if isCorrect else 'Incorrect Answer!'
        modalMessage = f'Your choice: {selectedWord}' if not isCorrect else ''
        selectedWord = f"Your choice:  {selectedWord}"
        
        # Render the modal HTML using Jinja2 template
        modal_html = render_template(
            'learning/modals/definition-pop-up.html',
            modalHeaderClass=modalHeaderClass,
            modalTitle=modalTitle,
            modalMessage=modalMessage,
            correctWord=translation,
            fullSentence=definition,
            exampleTranslation=word,
            theme=theme
        )

        return jsonify({'modal_html': modal_html})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@user_route.route('/modals/image-pop-up', methods=['GET'])
@login_required
def image_pop_up():
    # print(os.path.abspath('templates/learning/modals/pop-up.html'))
    # Pobierz parametry modalnego okienka
    selectedWord = request.args.get('selectedWord')
    correctWord = request.args.get('correctWord')
    theme = request.args.get('theme', 'light')
    modalHeaderClass = 'bg-success text-white' if selectedWord == correctWord else 'bg-danger text-white'
    modalTitle = 'Poprawna odpowiedź!' if selectedWord == correctWord else 'Nieprawidłowa odpowiedź!'
    modalMessage = modalTitle
    
    # Renderuj szablon modalnego okienka
    modal_html = render_template(
        'learning/modals/image-pop-up.html',
        modalHeaderClass=modalHeaderClass,
        modalTitle=modalTitle,
        modalMessage=modalMessage,
        correctWord=correctWord,
        theme=theme
    )
    return jsonify({'modal_html': modal_html})

@user_route.route('/modals/insert-pop-up', methods=['GET'])
@login_required
def insert_pop_up():
    selectedWord = request.args.get('selectedWord')
    correctWord = request.args.get('correctWord')
    fullSentence = request.args.get('fullSentence')
    exampleTranslation = request.args.get('exampleTranslation')
    theme = request.args.get('theme', 'light')

    modalHeaderClass = 'bg-success text-white' if selectedWord == correctWord else 'bg-danger text-white'
    modalTitle = 'Correct Answer!' if selectedWord == correctWord else 'Incorrect Answer!'
    modalMessage = f'Your choise: {selectedWord}' if selectedWord != correctWord else ''

    modal_html = render_template(
        'learning/modals/insert-pop-up.html',
        modalHeaderClass=modalHeaderClass,
        modalTitle=modalTitle,
        modalMessage=modalMessage,
        correctWord=correctWord,
        fullSentence=fullSentence,
        exampleTranslation=exampleTranslation,
        theme=theme
    )
    return jsonify({'modal_html': modal_html})

@user_route.route('/modals/multi-pop-up', methods=['GET'])
@login_required
def multi_pop_up():
    # Pobierz parametry modalnego okienka
    userTranslation = request.args.get('userTranslation')
    correctTranslation = request.args.get('correctTranslation')
    theme = request.args.get('theme', 'light')
    isCorrect = userTranslation.lower() == correctTranslation.lower()
    modalHeaderClass = 'bg-success text-white' if isCorrect else 'bg-danger text-white'
    modalTitle = 'Poprawna odpowiedź!' if isCorrect else 'Nieprawidłowa odpowiedź!'
    modalMessage = modalTitle
    correctWordMessage = correctTranslation
    fullSentenceMessage = ''  # Możesz tu dodać pełne zdanie, jeśli jest dostępne
    sentenceTranslation = ''  # Możesz tu dodać tłumaczenie zdania, jeśli jest dostępne

    # Renderuj szablon modalnego okienka
    modal_html = render_template(
        'learning/modals/multi-pop-up.html',
        modalHeaderClass=modalHeaderClass,
        modalTitle=modalTitle,
        modalMessage=modalMessage,
        correctTranslation=correctWordMessage,
        fullSentenceMessage=fullSentenceMessage,
        sentenceTranslation=sentenceTranslation,
        theme=theme
    )
    return jsonify({'modal_html': modal_html})

# funkcja nie działa w pełni offline pomimo pobrania repozytoriów whisper do folderu models
@user_route.route('/real-time-speech-recognition', methods=['POST'])
@login_required
def real_time_speech_recognition():

    # Load the model and processor
    processor = WhisperProcessor.from_pretrained(WHISPER_MEDIUM, local_files_only=True)
    model = WhisperForConditionalGeneration.from_pretrained(WHISPER_MEDIUM, local_files_only=True)

    try:
        CHUNK = 1024
        FORMAT = pyaudio.paInt16
        CHANNELS = 1
        RATE = 16000

        p = pyaudio.PyAudio()
        stream = p.open(format=FORMAT,
                        channels=CHANNELS,
                        rate=RATE,
                        input=True,
                        frames_per_buffer=CHUNK)

        print("Rozpoczęto nagrywanie...")

        frames = []

        while True:
            data = stream.read(CHUNK)
            frames.append(data)
            audio_data = np.frombuffer(data, dtype=np.int16)

            inputs = processor(audio_data, return_tensors="pt", sampling_rate=RATE)

            with torch.no_grad():
                generated_ids = model.generate(inputs.input_features)
                transcription = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]

            print("Transkrypcja: ", transcription)
            # Send partial transcription here if needed

        return jsonify({'transcription': transcription})

    except KeyboardInterrupt:
        print("Zatrzymano nagrywanie.")
        stream.stop_stream()
        stream.close()
        p.terminate()
        return jsonify({'message': 'Recording stopped'}), 200

    except Exception as e:
        print(f"Error in real-time speech recognition: {e}")
        return jsonify({'error': 'Failed to process speech'}), 500

    finally:
        stream.stop_stream()
        stream.close()
        p.terminate()

@user_route.route('/<path:path>')
@login_required
def send_static(path):
    return send_from_directory('public', path)

@user_route.route('/audio_files/English/<path:path>')
@login_required
def send_audio(path):
    return send_from_directory(AUDIO_FOLDER, path)

@user_route.route('/text-to-speech', methods=['POST'])
@login_required
def text_to_speech():
    print("Received request to convert text to speech")
    text = request.json.get('text', '')
    text = text.lower().replace('/', ' ') 
    print(f"Received text: {text}")
    if not text:
        print("Error: No text provided")
        return jsonify({'error': 'No text provided'}), 400
    
    try:
        mp3_path = os.path.join(AUDIO_FOLDER, f"{text}.mp3")
        if not os.path.exists(mp3_path):
            print("Generating audio...")

            if torch.cuda.is_available():
                # Run on GPU with cuda, required 12 GB vram
                # Zgodne z CUDA 12.0 / pytorch 2.0+  // Układy GPU Ampere, Ada lub Hopper (np. A100, RTX 3090, RTX 4090, H100). Wsparcie dla Turing GPU (T4, RTX 2080) już wkrótce, proszę użyć FlashAttention 1.x dla Turing GPU na razie.
                processor = AutoProcessor.from_pretrained(BARK_MODEL)
                # python -m pip install git+https://github.com/huggingface/optimum.git
                # pip install -U flash-attn --no-build-isolation
                
                device = "cuda" if torch.cuda.is_available() else "cpu"
                model = BarkModel.from_pretrained(BARK_MODEL, torch_dtype=torch.float16, attn_implementation="flash_attention_2").to(device)
            
                # enable CPU offload / włącz dodatkowo obciążenie CPU
                model.enable_cpu_offload()
            else:
                # Run with CPU
                processor = AutoProcessor.from_pretrained(BARK_MODEL)
                model = AutoModel.from_pretrained(BARK_MODEL)
            
            # voice_preset = "v2/en_speaker_6"

            inputs = processor(
                text,
                return_tensors="pt",
                # voice_preset=voice_preset,
            )

            speech_values = model.generate(**inputs, do_sample=True)

            sampling_rate = 22050
            
            wav_data = speech_values.cpu().numpy().squeeze()
            wav_path = os.path.join(AUDIO_FOLDER, f"{text}.wav")
            scipy.io.wavfile.write(wav_path, rate=sampling_rate, data=wav_data)
            
            audio_segment = AudioSegment.from_wav(wav_path)
            audio_segment.export(mp3_path, format="mp3")
            
            # Remove the intermediate WAV file after exporting to MP3
            if os.path.exists(wav_path):
                os.remove(wav_path)
                print(f"Removed intermediate WAV file: {wav_path}")
            
            print(f"Saving audio to file: {mp3_path}")
        else:
            print(f"Audio file already exists: {mp3_path}")
            
        return jsonify({'audio_path': mp3_path})
    except Exception as e:
        print(f"Error in text-to-speech: {e}")
        logging.error(f"Error in text-to-speech: {e}")
        return jsonify({'error': 'Failed to generate speech'}), 500
    
@user_route.route('/text-to-speech-groq', methods=['POST'])
@login_required
def text_to_speech_groq():

    API_KEY = "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"
    client = Groq(api_key=API_KEY)

    print("Received request to convert text to speech using Groq API")
    text = request.json.get('text', '')
    text = text.lower().replace('/', ' ')
    print(f"Received text: {text}")
    if not text:
        print("Error: No text provided")
        return jsonify({'error': 'No text provided'}), 400

    flaga = 0 # groq nie obsługuje text-to-speech tylko speech to text
    try:
        mp3_path = os.path.join(AUDIO_FOLDER, f"{text}.mp3")
        if not os.path.exists(mp3_path) and flaga == 1:
            print("Generating audio...")
            completion = client.chat.completions.create(
                model="whisper-large-v3",
                messages=[{"role": "user", "content": text}],
                temperature=1,
                max_tokens=1024,
                top_p=1,
                stream=True,
                stop=None,
            )

            audio_content = b""
            for chunk in completion:
                audio_content += chunk.choices[0].delta.audio or b""

            audio_path = os.path.join(AUDIO_FOLDER, f"{text}.mp3")
            with open(audio_path, 'wb') as audio_file:
                audio_file.write(audio_content)

            print("Plik został zapisany pod ścieżką: ", audio_path)
    
        else:
            print(f"Audio file already exists: {mp3_path}")
            
        return jsonify({'audio_path': mp3_path})
        
    except Exception as e:
        print(f"Error in text-to-speech using Groq: {e}")
        logging.error(f"Error in text-to-speech using Groq: {e}")
        return jsonify({'error': 'Failed to generate speech using Groq'}), 500

@user_route.route('/load-audio-paths', methods=['POST'])
@login_required
def load_audio_paths():
    words = request.json['words']
    audio_paths = []
    for word in words:
        filename = word.lower().replace(' ', '-')
        audio_path = os.path.join(AUDIO_FOLDER, f'{filename}.mp3').lower()
        if os.path.exists(audio_path):
            log_to_file(audio_logger, f'File already exists: {audio_path}')
            audio_paths.append(audio_path)
        else:
            url = f'https://www.ang.pl/sound/dict/{filename}.mp3'
            response = requests.get(url, stream=True)
            if response.status_code == 200:
                with open(audio_path, 'wb') as f:
                    for chunk in response.iter_content(1024):
                        f.write(chunk)
                log_to_file(audio_logger, f'Downloaded audio: {url} to {audio_path}')
                audio_paths.append(audio_path)
            else:
                filename = filename.replace('-', '_')
                url = f'https://www.diki.pl/images-common/en/mp3/{filename}.mp3'
                response = requests.get(url, stream=True)
                if response.status_code == 200:
                    with open(audio_path, 'wb') as f:
                        for chunk in response.iter_content(1024):
                            f.write(chunk)
                    log_to_file(audio_logger, f'Downloaded audio: {url} to {audio_path}')
                    audio_paths.append(audio_path)
                else:
                    log_to_file(audio_logger, f'Failed to download audio from: {url}')
                    audio_paths.append(None)
    log_to_file(app_logger, f'Loaded audio paths: {audio_paths}') 
    return jsonify(audio_paths)

def log_to_file(logger, message):
    logger.info(message)

def check_and_download_file(url, file_path, logger):
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # Raise HTTPError for bad responses (4xx and 5xx)
        with open(file_path, 'wb') as file:
            file.write(response.content)
        return file_path
    except requests.exceptions.RequestException as e:
        log_to_file(logger, f'Error downloading {url}: {e}')
        return None

@user_route.route('/load-image-paths', methods=['POST'])
@login_required
def load_image_paths():
    try:
        data = request.get_json()
        words = data.get('words', [])

        image_paths = []
        for word in words:
            if not word:
                image_paths.append(None)
                continue

            filename = word.lower().replace(' ', '-')
            image_path = os.path.join(IMAGE_FOLDER, f"{filename}.jpg")

            if os.path.exists(image_path):
                log_to_file(image_logger, f'File already exists: {image_path}')
                image_paths.append(image_path)
                continue

            url = f"https://www.ang.pl/img/slownik/{filename}.jpg"
            image_file = check_and_download_file(url, image_path, image_logger)

            if image_file:
                log_to_file(image_logger, f'Downloaded image: {url} to {image_path}')
                image_paths.append(image_file)
            else:
                log_to_file(image_logger, f'Failed to download image from: {url}')
                image_paths.append(None)

        return jsonify(image_paths)

    except Exception as e:
        log_to_file(app_logger, f'Error in load_image_paths: {e}')
        return jsonify({"error": "Internal Server Error"}), 500

@user_route.route('/save', methods=['POST'])
@login_required
def save_json():
    file_name = request.json['fileName']
    json_data = request.json['jsonData']
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    with open(file_path, 'w') as f:
        json.dump(json_data, f, indent=2)
    print(f'Saved JSON file: {file_path}')  # Print detailed information
    return jsonify({'message': 'Data saved successfully as JSON'})

@user_route.route('/save_statistic', methods=['POST'])
@login_required
def save_statistic():
    statistic = request.json
    logging.info(f'Received statistic: {statistic}')
    if not os.path.exists(STATISTICS_FILE):
        with open(STATISTICS_FILE, 'w') as f:
            json.dump([statistic], f, indent=2)
    else:
        with open(STATISTICS_FILE, 'r+') as f:
            data = f.read()
            if data:
                existing_data = json.loads(data)
            else:
                existing_data = []
            existing_data.append(statistic)
            f.seek(0)
            json.dump(existing_data, f, indent=2)
    print(f'Saved statistic: {statistic}') 
    return jsonify({'message': 'Statistic saved successfully'})

@user_route.route('/saveSetting', methods=['POST'])
@login_required
def save_setting():
    data = request.json['excludedWords']
    with open(SETTING_FILE, 'w') as f:
        json.dump(data, f, indent=2)
    print(f'Saved setting: {data}')  
    return jsonify({'message': 'Settings update requested'})

@user_route.route('/words', methods=['GET'])
@login_required
def get_words():
    file_name = request.args.get('file')
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            data = json.load(f)
            return jsonify(data)
    else:
        return jsonify([])

@user_route.route('/edit/:file/:index', methods=['POST'])
@login_required
def edit_word():
    file_name = request.args.get('file')
    index = int(request.args.get('index'))
    updated_data = request.json
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    if os.path.exists(file_path):
        with open(file_path, 'r+') as f:
            data = json.load(f)
            if index >= 0 and index < len(data):
                data[index] = updated_data
                f.seek(0)
                json.dump(data, f, indent=2)
                return jsonify({'message': 'Data updated successfully'})
            else:
                return jsonify({'error': 'Invalid index'})
    else:
        return jsonify({'error': 'File not found'})

@user_route.route('/files', methods=['GET'])
@login_required
def get_files():
    files = os.listdir(UPLOAD_FOLDER)
    return jsonify(files)

@user_route.route('/upload', methods=['POST'])
@login_required
def upload_file():
    file = request.files['file']
    file_path = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))
    file.save(file_path)
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return jsonify({'columns': data[0]})

data = [
    {
        "language": "English",
        "translationLanguage": "Polish",
        "word": "Include",
        "translation": "zawierać",
        "definition": "To contain something as a part of something else.",
        "example": "The report includes several recommendations.",
        "example_translation": "Raport zawiera kilka zaleceń.",
        "imageLink": "image_files/English/include.jpg",
        "audioLink": "audio_files/English/include.mp3"
    }
    # Dodaj więcej obiektów w miarę potrzeby
]

@user_route.route('/data', methods=['GET'])
@login_required
def get_data():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 5))
    start = (page - 1) * per_page
    end = start + per_page
    return jsonify(data[start:end])

@user_route.route('/data', methods=['POST'])
@login_required
def add_data():
    new_item = request.json
    data.append(new_item)
    return jsonify(data)

@user_route.route('/data/<int:index>', methods=['PUT'])
@login_required
def update_data(index):
    updated_item = request.json
    data[index] = updated_item
    return jsonify(data)

@user_route.route('/data/<int:index>', methods=['DELETE'])
@login_required
def delete_data(index):
    del data[index]
    return jsonify(data)

# Globalna zmienna do przechowywania danych
data = []

@user_route.route('/parse_excel_columns', methods=['POST'])
@login_required
def parse_excel_columns():
    file = request.files['file']
    df = pd.read_excel(file)
    columns = df.columns.tolist()
    return jsonify(columns=columns)

@user_route.route('/upload_excel', methods=['POST'])
@login_required
def upload_excel():
    file = request.files['file']

    # Retrieve column names from the form, use .get() to handle empty keys
    language_col = request.form.get('language', '')
    translationLanguage_col = request.form.get('translationLanguage', '')
    word_col = request.form.get('word', '')
    translation_col = request.form.get('translation', '')
    definition_col = request.form.get('definition', '')
    example_col = request.form.get('example', '')
    example_translation_col = request.form.get('example_translation', '')
    imageLink_col = request.form.get('imageLink', '')
    audioLink_col = request.form.get('audioLink', '')

    # Read the Excel file
    df = pd.read_excel(file)
    print(df.columns.tolist())

    for _, row in df.iterrows():
        entry = {
            "language": row[language_col] if language_col and pd.notna(row[language_col]) else '',
            "translationLanguage": row[translationLanguage_col] if translationLanguage_col and pd.notna(row[translationLanguage_col]) else '',
            "word": row[word_col] if word_col and pd.notna(row[word_col]) else '',
            "translation": row[translation_col] if translation_col and pd.notna(row[translation_col]) else '',
            "definition": row[definition_col] if definition_col and pd.notna(row[definition_col]) else '',
            "example": row[example_col] if example_col and pd.notna(row[example_col]) else '',
            "example_translation": row[example_translation_col] if example_translation_col and pd.notna(row[example_translation_col]) else '',
            "imageLink": row[imageLink_col] if imageLink_col and pd.notna(row[imageLink_col]) else '',
            "audioLink": row[audioLink_col] if audioLink_col and pd.notna(row[audioLink_col]) else ''
        }
        data.append(entry)
    
    return jsonify(data)


@user_route.route('/download_configuration', methods=['POST'])
@login_required
def download_configuration():
    try:
        request_data = request.get_json()
        columns = request_data.get('columns')
        rows = request_data.get('rows')
        format = request_data.get('format')

        if not columns:
            return jsonify({'error': 'No columns selected'}), 400
        
        if not rows:
            return jsonify({'error': 'No rows selected'}), 400

        if not data:
            return jsonify({'error': 'No data available to download'}), 400
        
        if format not in ['json', 'csv', 'excel']:
            return jsonify({'error': 'Unsupported format'}), 400
        
        # Filter data based on selected columns and rows
        filtered_data = [{col: row[col] for col in columns} for i, row in enumerate(data) if str(i) in rows]

        if format == 'json':
            response = jsonify(filtered_data)
            response.headers['Content-Disposition'] = 'attachment; filename=configuration.json'
            response.mimetype = 'application/json'
            return response
        elif format == 'csv':
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=columns)
            writer.writeheader()
            writer.writerows(filtered_data)
            output.seek(0)
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = 'attachment; filename=configuration.csv'
            response.mimetype = 'text/csv'
            return response
        elif format == 'excel':
            output = BytesIO()
            df = pd.DataFrame(filtered_data)
            df.to_excel(output, index=False)
            output.seek(0)
            response = make_response(output.getvalue())
            response.headers['Content-Disposition'] = 'attachment; filename=configuration.xlsx'
            response.mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@user_route.route('/upload_json', methods=['POST'])
@login_required
def upload_json():
    file = request.files['file']
    global data
    data = json.load(file)
    return jsonify(data)

@user_route.route('/data/count', methods=['GET'])
@login_required
def get_data_count():
    print("len data/count", len(data))
    return jsonify(len(data))

# Endpoint do zapisywania JSON, służy do zapisywania całych zestawów, które zostały poprawione w ekselu wbudowanym w aplikacje
@user_route.route('/uploads-save-json', methods=['POST'])
@login_required
def uploads_save_json():
    data = request.get_json()
    file_name = data.get('fileName', 'data')  ## w chwili obecnej zmienna fileName nie jest przekazywana z frontendu
    json_data = data.get('data')   
    print(file_name)
    try:
        file_path = os.path.join(UPLOAD_FOLDER, f"{file_name}.json")
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(json_data)
        # Logowanie zapisu pliku
        upload_logger.info(f"Plik zapisany: {file_path}")
        return jsonify(success=True, file_path=file_path)
    except Exception as e:
        # Logowanie błędu
        upload_logger.error(f"Błąd podczas zapisywania pliku: {e}")
        return jsonify(success=False)

@user_route.route('/upload-save-json', methods=['POST'])
@login_required
def upload_save_json():
    global data
    data = request.json
    with open('data.json', 'w') as f:
        json.dump(data, f)
    return jsonify({"success": True})

@user_route.route('/translate-word', methods=['GET'])
@login_required
def translate_word():
    word = request.args.get('word')
    target_lang = request.args.get('targetLang')
    # translation = translate(word, target_lang)
    # return jsonify(translation)